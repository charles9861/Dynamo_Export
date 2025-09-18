import clr
import os

# Add Assemblies
clr.AddReference('AcMgd')
clr.AddReference('AcCoreMgd')
clr.AddReference('AcDbMgd')
clr.AddReference('AeccDbMgd')

# Civil 3D
from Autodesk.AutoCAD.ApplicationServices import *
from Autodesk.Civil.DatabaseServices import *

# Excel
import openpyxl
from openpyxl import Workbook

# IN[0] = list of PipeNetwork objects from Dynamo
networks = IN[0]

# Output Excel path
excel_path = r"C:\Data Extract\PipeNetworkData.xlsx"

pipes_data = []
structures_data = []
parts_data = set()

# Loop networks
for net in networks:
    # --- Pipes ---
    for pipe in net.Pipes:
        pipes_data.append([
            net.Name,
            pipe.Handle.ToString(),
            pipe.Name,
            pipe.Length,
            getattr(pipe, "InnerDiameter", ""),
            getattr(pipe, "OuterDiameter", ""),
            getattr(pipe, "Slope", ""),
            getattr(pipe, "StartInvertElevation", ""),
            getattr(pipe, "EndInvertElevation", "")
        ])

    # --- Structures ---
    for struct in net.Structures:
        rim = getattr(struct, "RimElevation", "")
        sump = getattr(struct, "SumpElevation", "")
        sump_depth = rim - sump if rim and sump else ""
        part_family, part_size = "", ""
        try:
            part = struct.Part
            part_family = getattr(part, "FamilyName", "")
            part_size = getattr(part, "SizeName", "")
            parts_data.add((part_family, part_size))
        except:
            pass

        structures_data.append([
            net.Name,
            struct.Handle.ToString(),
            struct.Name,
            part_family,
            part_size,
            rim,
            sump,
            sump_depth
        ])

# --- Excel Export ---
if os.path.exists(excel_path):
    try:
        os.remove(excel_path)
    except:
        raise Exception("File exists and cannot be deleted. Close it if it's open in Excel.")

wb = Workbook()

# Pipes sheet
ws1 = wb.active
ws1.title = "Pipes"
ws1.append(["Network", "Handle", "Name", "Length", "InnerDiameter", "OuterDiameter", "Slope", "StartInvert", "EndInvert"])
for row in pipes_data:
    ws1.append(row)

# Structures sheet
ws2 = wb.create_sheet(title="Structures")
ws2.append(["Network", "Handle", "Name", "PartFamily", "PartSize", "RimElevation", "SumpElevation", "SumpDepth"])
for row in structures_data:
    ws2.append(row)

# Parts sheet
ws3 = wb.create_sheet(title="Parts")
ws3.append(["PartFamily", "PartSize"])
for fam, size in sorted(parts_data):
    ws3.append([fam, size])

wb.save(excel_path)

OUT = "Excel export complete: " + excel_path
