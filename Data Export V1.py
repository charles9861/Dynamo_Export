# Load the Python Standard and DesignScript Libraries
import sys
import clr
import os

# Add Assemblies for AutoCAD and Civil3D
clr.AddReference('AcMgd')
clr.AddReference('AcCoreMgd')
clr.AddReference('AcDbMgd')
clr.AddReference('AecBaseMgd')
clr.AddReference('AecPropDataMgd')
clr.AddReference('AeccDbMgd')

# Import references from AutoCAD
from Autodesk.AutoCAD.ApplicationServices import Application
from Autodesk.AutoCAD.DatabaseServices import *
from Autodesk.AutoCAD.Geometry import *

# Import references from Civil3D
from Autodesk.Civil.ApplicationServices import *
from Autodesk.Civil.DatabaseServices import *

# Try importing openpyxl (must be available in Dynamo's Python environment)
try:
    import openpyxl
    from openpyxl import Workbook
except:
    raise Exception("openpyxl is required for Excel export. Install into Dynamo's Python environment.")

# Path to Excel file
excel_path = r"C:\Data Extract\PipeNetworkData.xlsx"

# The inputs to this node will be stored as a list in the IN variables.
dataEnteringNode = IN

adoc = Application.DocumentManager.MdiActiveDocument

pipes_data = []
structures_data = []
parts_data = set()

with adoc.LockDocument():
    db = adoc.Database
    tm = db.TransactionManager
    with tm.StartTransaction() as t:
        civdoc = CivilApplication.ActiveDocument
        net_ids = civdoc.GetPipeNetworkIds()

        for net_id in net_ids:
            net = t.GetObject(net_id, OpenMode.ForRead)

            # --- Pipes ---
            try:
                pipes = net.GetPipes()
            except:
                pipes = []

            for p in pipes:
                pipe = t.GetObject(p, OpenMode.ForRead)
                pipes_data.append([
                    net.Name,
                    pipe.Handle.ToString(),
                    pipe.Name,
                    pipe.Length,
                    pipe.InnerDiameter if hasattr(pipe, "InnerDiameter") else "",
                    pipe.OuterDiameter if hasattr(pipe, "OuterDiameter") else "",
                    pipe.Slope if hasattr(pipe, "Slope") else "",
                    pipe.StartInvertElevation if hasattr(pipe, "StartInvertElevation") else "",
                    pipe.EndInvertElevation if hasattr(pipe, "EndInvertElevation") else ""
                ])

            # --- Structures ---
            try:
                structs = net.GetStructures()
            except:
                structs = []

            for s in structs:
                struct = t.GetObject(s, OpenMode.ForRead)
                rim = struct.RimElevation if hasattr(struct, "RimElevation") else ""
                sump = struct.SumpElevation if hasattr(struct, "SumpElevation") else ""
                sump_depth = rim - sump if rim and sump else ""
                part_family = ""
                part_size = ""
                try:
                    part = struct.Part
                    part_family = part.FamilyName if hasattr(part, "FamilyName") else ""
                    part_size = part.SizeName if hasattr(part, "SizeName") else ""
                    parts_data.add((part_family, part_size))
                except:
                    pass

                structures_data.append([
                    net.Name,
                    struct.Handle.ToString(),
                    struct.Name,
                    part_family,
                    part_size,
                    rim,
                    sump,
                    sump_depth
                ])

        t.Commit()

# --- Write Excel ---
wb = Workbook()

# Pipes sheet
ws_pipes = wb.active
ws_pipes.title = "Pipes"
ws_pipes.append(["Network", "Handle", "Name", "Length", "InnerDiameter", "OuterDiameter", "Slope", "StartInvert", "EndInvert"])
for row in pipes_data:
    ws_pipes.append(row)

# Structures sheet
ws_struct = wb.create_sheet("Structures")
ws_struct.append(["Network", "Handle", "Name", "PartFamily", "PartSize", "RimElevation", "SumpElevation", "SumpDepth"])
for row in structures_data:
    ws_struct.append(row)

# Parts list sheet
ws_parts = wb.create_sheet("PartsList")
ws_parts.append(["PartFamily", "PartSize"])
for fam, size in sorted(parts_data):
    ws_parts.append([fam, size])

# Save workbook
wb.save(excel_path)

OUT = "Excel file created at: " + excel_path
